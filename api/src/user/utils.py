import json
import os
import uuid
import tempfile
from fastapi import Response, HTTPException
from passlib.context import CryptContext
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from fastapi.encoders import jsonable_encoder
from pydantic import EmailStr
from starlette import status
from .models import User
from .schemas import UserIn
from helpers import Utils
from worker import celery


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


class AfterCreateUser:
    @staticmethod
    async def send_email_noti(user: User, not_hash_password: str):
        await Utils.send_email(user, not_hash_password)


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)

def create_temp_file():
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    with os.fdopen(fd, 'w') as f:
        f.write('TEST\n')
    try:
        yield path
    finally:
        os.unlink(path)

def remove_file(path: str) -> None:
    os.unlink(path)

@celery.task
def download_excel(queryset: list):
    wb = Workbook()
    ws = wb.active

    headings = [
        "STT",
        "Username",
        "Email",
        "Fullname",
        "Is admin",
        "Is active",
    ]

    for column, heading in enumerate(headings, 1):
        ws.cell(row=1, column=column).value = heading
    for row, user in enumerate(queryset, 2):
        values = [
            row - 1,
            user["username"],
            user["email"],
            user["full_name"],
            "admin" if user["is_admin"] else "normal",
            "active" if user["is_active"] else "inactive",
        ]
        for column, value in enumerate(values, 1):
            ws.cell(row=row, column=column).value = value

    file_name = "list-user.xlsx"
    file_path = excel_dest(file_name)
    wb.save(file_path)

    return file_path


async def import_user(sheet, user_collections, result: list):
    col_map = get_column_map()
    start_row = 1
    for index, row in enumerate(sheet):
        if index < start_row:
            continue
        c = get_cell_data(row)
        to_str_inner = to_str(c, col_map)
        data = {
            "username": to_str_inner("username") or None,
            "full_name": to_str_inner("full_name"),
            "email": to_str_inner("email"),
            "password": to_str_inner("password"),
            "is_admin": True if to_str_inner("is_admin") == "admin" else False,
        }
        user = await create(UserIn(**data), user_collections)
        result.append(user)
    return result


def image_dest(filename):
    ext = filename.split(".")[-1]
    return os.path.join("./static/images", f"{uuid.uuid4()}.{ext}")


def excel_dest(filename):
    ext = filename.split(".")[-1]
    return os.path.join("./static/excel", f"{uuid.uuid4()}.{ext}")


def get_column_map():
    return {
        "username": 1,
        "email": 2,
        "full_name": 3,
        "password": 4,
        "is_admin": 5,
    }


def get_cell_data(row, start_column=0):
    def inner(index):
        value = row[start_column + index]
        if isinstance(value, str):
            value = value.strip()
        return value

    return inner


def to_str(c, col_map):
    def inner(field):
        return "" if c(col_map[field]) is None else str(c(col_map[field])).strip()

    return inner


async def create(payload: UserIn, user_collections):
    user = await user_collections.find_one(
        {
            "$or": [
                {"username": payload.username.lower()},
                {"email": payload.email.lower()},
            ]
        },
    )
    if user:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail="Account already exist"
        )
    not_hash_password = payload.password
    payload.password = get_password_hash(payload.password)
    payload.email = EmailStr(payload.email.lower())
    payload.is_active = True

    new_user = User(**payload.dict())

    new_user_dict = jsonable_encoder(new_user)
    user_collections.insert_one(new_user_dict)
    await AfterCreateUser.send_email_noti(new_user, not_hash_password)
    return new_user
